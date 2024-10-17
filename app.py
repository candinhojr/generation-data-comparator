import io
import json
import openpyxl
import pandas as pd
import streamlit as st
from streamlit_modal import Modal
from openpyxl.formatting.rule import CellIsRule
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill

# CONSTRUÇÃO DE CLASSES E MÉTODOS PARA COLETA DOS DADOS DA PLANILHA
class ReadData:
    def __init__(self, path: str) -> None:
        self.path = path

    def read_ger_data(self, santa_monica: bool = False) -> pd.DataFrame:
        """
        Método responsável pela leitura da geração verificada (aplicado aos dados do PIM e do SAGER).
        santa_monica = True indica que os dados de Santa Mônica e Trairí são expostos separadamente (PIM).
        santa_monica = False indica que os dados de Santa Mônica e Trairí são expostos juntos (SAGER).
        """
        data = pd.read_excel(self.path, header=None, engine="openpyxl")
        data_index = data.iloc[1:, 0]
        data_columns = [str(col).split(':')[0] for col in data.iloc[0, 1:]]
        data_values = data.iloc[1:, 1:]
        data_df = pd.DataFrame(data=data_values.values, index=data_index.values, columns=data_columns)
        if santa_monica:
            data_df['CETR'] = data_df['CETR'] + data_df['CESM']
            data_df = data_df.drop(columns=['CESM'])
        return data_df

# CONSTRUÇÃO DE CLASSES E MÉTODOS PARA ORGANIZAÇÃO DOS DADOS COLETADOS DA PLANILHA
class CompareData:

    @staticmethod
    def compare_ger_data(pim_data: pd.DataFrame, sager_ger_data: pd.DataFrame, tol_ger: float = 3):
        """
        Método utilizado para comparar os dados de Geração Verificada do SIGO e do SAGER.
        """
        compare = pd.DataFrame()
        for column in pim_data.columns:
            compare[f'{column} PIM'] = pim_data[column]
            compare[f'{column} SAGER'] = sager_ger_data[column].values
            compare_status = []
            for idx in range(len(pim_data[column])):
                if abs(pim_data[column][idx] - sager_ger_data[column][idx]) <= tol_ger:
                    compare_status.append('Dado Correto')
                else:
                    compare_status.append(round(pim_data[column][idx], 3))
            compare[f'{column} STATUS'] = compare_status
        return compare[sorted(compare.columns)]

    @staticmethod
    def create_df_per_index(compare: dict, index: list) -> pd.DataFrame:
        """
        Método auxiliar para separar o dataframe criado em compare_coff_data em vários dataframes, baseados no índice (sigla da Usina).
        """
        columns = [
            f'{index} HORA-INÍCIO STATUS',
            f'{index} HORA-FIM STATUS',
            f'{index} RAZÃO STATUS',
            f'{index} ORIGEM STATUS',
            f'{index} LIMITAÇÃO STATUS',
            f'{index} HORA-INÍCIO SIGO',
            f'{index} HORA-INÍCIO SAGER',
            f'{index} HORA-FIM SIGO',
            f'{index} HORA-FIM SAGER',
            f'{index} RAZÃO SIGO',
            f'{index} RAZÃO SAGER',
            f'{index} ORIGEM SIGO',
            f'{index} ORIGEM SAGER',
            f'{index} LIMITAÇÃO SIGO',
            f'{index} LIMITAÇÃO SAGER']
        columns = [col for col in columns if col in compare]
        return pd.DataFrame({col: compare[col] for col in columns})

def generate_excel(df_ger, tabs_list):
    output = io.BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for sheet_name in tabs_list:
        if sheet_name in df_ger:
            ws = workbook.create_sheet(title=sheet_name)
            df = df_ger[sheet_name]
            
            # Write column headers
            for c_idx, column in enumerate(df.columns, 2):  # Start from column B
                cell = ws.cell(row=1, column=c_idx, value=column)
                cell.font = Font(bold=True)
                cell.border = thin_border
            
            # Write data rows
            for r_idx, (index, row) in enumerate(df.iterrows(), 2):  # Start from row 2
                # Write index (date) in column A
                cell = ws.cell(row=r_idx, column=1, value=index)
                cell.font = Font(bold=True)
                cell.border = thin_border
                
                # Write row data
                for c_idx, value in enumerate(row, 2):  # Start from column B
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Apply conditional formatting to status column
                    if df.columns[c_idx-2] == f'{sheet_name} STATUS':
                        if value == "Dado Correto":
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        elif isinstance(value, (int, float)):
                            cell.fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')

    workbook.save(output)
    output.seek(0)
    return output

# CONSTRUÇÃO DA LÓGICA PRINCIPAL DO PROGRAMA
# Configurar a página
st.set_page_config(layout="wide", page_title="Geração", page_icon=":zap:")
st.title("GERAÇÃO VERIFICADA")

# Configurar modais
info_modal = Modal(key="-INFO_MODAL-", title="INFORMAÇÕES")
config_modal = Modal(key="-CONFIG_MODAL-", title="CONFIGURAÇÕES")

if 'index_list' not in st.session_state:
    st.session_state['index_list'] = json.load(open('default.json', 'r'))

index_list = st.session_state['index_list'].split(",")

# Configurar botões de colunas
col1, col2, col3, col4, col5, col6 = st.columns(6)
with col1:
    if st.button("⚠️ INFO."):
        info_modal.open()
with col2:
    if st.button("⚙️ CONFIG."):
        config_modal.open()
with col3:
    buffer_geracao_verificada = io.BytesIO()
    with pd.ExcelWriter(buffer_geracao_verificada, engine='xlsxwriter') as writer:
        pd.DataFrame(columns=index_list, index=['00:30', '01:00', '01:30', '02:00', '02:30', '03:00', '03:30', '04:00', '04:30', '05:00', '05:30', '06:00',
                                                '06:30', '07:00', '07:30', '08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00',
                                                '12:30', '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00',
                                                '18:30', '19:00', '19:30', '20:00', '20:30', '21:00', '21:30', '22:00', '22:30', '23:00', '23:30', '00:00']
                                                ).to_excel(writer, index=True, sheet_name='SAGER - GER. VERIFICADA')
    st.download_button(
        label='⬇️ Modelo SAGER',
        data=buffer_geracao_verificada.getvalue(),
        file_name='sager_geracao_verificada.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        key='-DONWLOAD_MODELO_GER-'
    )

st.markdown("---")

if info_modal.is_open():
    with info_modal.container():
        st.info("""
        - **Upload TELEMEDIÇÃO/PIM:** Em https://pimpro.ds55.local/GraficosHtml5 faça o download, em formato Excel (.xlsx) dos dados de Geração Líquida (Demanda Ativa Del; Mega; 30min).\
                Selecione a mesma data inicial e final. Ao final, submeta o arquivo no campo indicado abaixo.\n
        - **Upload SAGER:** Em https://pops.ons.org.br/pop/#15883 (SAGER > Apuração de Renováveis > Apuração > Consistência do Agente), em "Patamares", copie os dados da coluna de "Geração Verificada (MWh/h)"\
                e cole-os em uma planilha Excel (.xlsx). Em "Restrições", copie todas as colunas e cole-as em outra planilha Excel (.xlsx). Isso é necessário tendo em vista que a ONS ainda não disponibilizou\
                a API para o download dos dados do SAGER. Para garantir o preenchimento da planilha no formato correto, faça o download dos modelos em "Modelo SAGER". Ao final, submeta o arquivo no campo\
                indicado abaixo.
        """)

if 'tol_ger' not in st.session_state:
    st.session_state['tol_ger'] = 3
if 'tol_min' not in st.session_state:
    st.session_state['tol_min'] = 6
if 'index_list' not in st.session_state:
    st.session_state['index_list'] = json.load(open('default.json', 'r'))
if config_modal.is_open():
    with config_modal.container():
        col1, col2 = st.columns(2)
        with col1:
            st.number_input("Tolerância [MW] para avaliação da Geração Verificada:", value=st.session_state['tol_ger'], key='-TOL_GER-')
        with col2:
            st.number_input("Tolerância [min] para as horas de registro de limitações:", value=st.session_state['tol_min'], key='-TOL_MIN-')
        st.session_state['index_list'] = st.text_area("Inclua/Exclua usinas para avaliação (separe por vírgula):", value=st.session_state['index_list'])

        if st.button("SALVAR"):
            st.session_state['tol_ger'] = float(st.session_state['-TOL_GER-'])
            st.session_state['tol_min'] = float(st.session_state['-TOL_MIN-'])
            json.dump(st.session_state['index_list'], open('default.json', 'w'))
            config_modal.close()
index_list = st.session_state['index_list'].split(",")

st.subheader("UPLOAD DOS ARQUIVOS")

col1, col2 = st.columns(2)

with col1:
    pim_ger_file = (st.file_uploader("**PIM:** Faça upload da planilha de Geração do TELEMEDIÇÃO/PIM em formato Excel (.xlsx)", type=["xlsx"]))
with col2:
    sager_ger_file = st.file_uploader("**SAGER:** Faça upload da planilha de Geração Verificada do SAGER em formato Excel (.xlsx)", type=["xlsx"])

st.markdown("---")

# Inicialize variáveis de estado
if 'df_ger' not in st.session_state:
    st.session_state.df_ger = None
if 'buffer_ger' not in st.session_state:
    st.session_state.buffer_ger = None

col1, col2, col3, col4, col5, col6, col7, col8 = st.columns(8)

with col1:
    execute_button = st.button("EXECUTAR", use_container_width=True)

with col8:
    if st.button("Limpar", use_container_width=True):
        st.session_state.df_ger = None
        st.session_state.buffer_ger = None
        st.experimental_rerun()

if execute_button:
    try:
        dado_correto = "Dado Correto"

        if not pim_ger_file:
            raise ValueError("Arquivo PIM não foi carregado.")
        if not sager_ger_file:
            raise ValueError("Arquivo SAGER não foi carregado.")
        
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # GERAÇÃO DO ARQUIVO DE VALIDAÇÃO DA GERAÇÃO VERIFICADA
        buffer_ger = io.BytesIO()

        try:
            pim_data = ReadData(pim_ger_file).read_ger_data(santa_monica=True)
        except Exception as e:
            raise ValueError(f"Erro ao ler dados do PIM: {str(e)}")
        
        try:
            sager_ger_data = ReadData(sager_ger_file).read_ger_data()
        except Exception as e:
            raise ValueError(f"Erro ao ler dados do SAGER: {str(e)}")
        
        try:
            compare_ger = CompareData.compare_ger_data(pim_data, sager_ger_data, st.session_state['tol_ger'])
        except Exception as e:
            raise ValueError(f"Erro ao comparar dados: {str(e)}")

        with pd.ExcelWriter(buffer_ger, engine='xlsxwriter') as writer:
            for idx in range(len(index_list)):
                compare_ger.iloc[:, (3*idx):(3*idx + 3)].to_excel(writer, index=True, sheet_name=sorted(index_list)[idx])
            writer._save()

        buffer_ger.seek(0)
        wb_ger = openpyxl.load_workbook(buffer_ger)
        for sheet_name in wb_ger.sheetnames:
            ws = wb_ger[sheet_name]
            for row in range(2, 50):
                ws.conditional_formatting.add(f"D{row}", CellIsRule(operator='equal', formula=['{dado_correto}'], stopIfTrue=True, fill=green_fill))
                ws.conditional_formatting.add(f"D{row}", CellIsRule(operator='notEqual', formula=['{dado_correto}'], stopIfTrue=True, fill=red_fill))

        buffer_ger = io.BytesIO()
        wb_ger.save(buffer_ger)
        buffer_ger.seek(0)
        df_ger = pd.read_excel(buffer_ger, sheet_name=None)

        # Rename 'Unnamed: 0' to 'Data' and set it as index
        for sheet_name in df_ger:
            # df_ger[sheet_name] = df_ger[sheet_name].rename(columns={'Unnamed: 0': 'Data'})
            # df_ger[sheet_name].set_index('Data', inplace=True)
            df_ger[sheet_name].set_index(df_ger[sheet_name].columns[0], inplace=True)
            df_ger[sheet_name].index.name = None  # Remove the index name
        
        st.session_state.df_ger = df_ger
        st.session_state.buffer_ger = buffer_ger

        st.success("Êxito. Agora você observar os resultados e ou efetuar o download do arquivo para análise.")

    except ValueError as e:
        st.error(f"Erro de execução: {str(e)}")
    except openpyxl.utils.exceptions.InvalidFileException:
        st.error("Erro: Arquivo Excel inválido. Verifique se os arquivos estão corrompidos ou no formato adequado.")
    except Exception as e:
        st.error(f"Erro inesperado: {str(e)}. Por favor, entre em contato com o suporte técnico.")

# O código para exibir as tabs e os dataframes fica aqui, fora do if
if st.session_state.df_ger is not None:
    tabs_list = st.session_state['index_list'].split(",")
    # Create tabs for each sheet
    tabs = st.tabs(tabs_list)

    # Function to highlight cells based on conditions
    def highlight_cells(val):
        if val == "Dado Correto":
            return 'background-color: #90EE90'  # Light green
        elif isinstance(val, (int, float)):
            return 'background-color: #FFCCCB'  # Light red
        else:
            return ''

    # Display each dataframe in its corresponding tab
    for tab, sheet_name in zip(tabs, st.session_state.df_ger.keys()):
        with tab:
            try:
                # df = df_ger[sheet_name]
                df = st.session_state.df_ger[sheet_name]

                # Apply conditional formatting
                styled_df = df.style.applymap(highlight_cells, subset=[f'{sheet_name} STATUS'])

                # Display the dataframe with custom formatting
                st.dataframe(
                    styled_df,
                    use_container_width=True,
                    height=500,  # Adjust this value as needed
                )
            except KeyError:
                st.warning(f"Dados para '{sheet_name}' não encontrados. Verifique se a usina está presente nos arquivos carregados.")

    excel_file = generate_excel(st.session_state.df_ger, tabs_list)
    st.download_button(
        label='⬇️ Geração Verificada',
        data=excel_file,
        file_name='geracao_verificada.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        key='-DONWLOAD_GERACAO_VERIFICADA-'
    )
    st.session_state.df_ger = None
    st.session_state.buffer_ger = None
